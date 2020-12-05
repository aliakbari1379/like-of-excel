
from tkinter import *
from tkinter import messagebox
from PIL import ImageTk,Image
import numpy as np
import pandas as pd
from tkinter import font as tkf
from openpyxl import *
from tkcalendar import *
from tkinter import ttk


root=Tk()
root.geometry('1700x980')
root.title("hiweb")
fonti= tkf.Font(family='Arial', size=11, weight=tkf.BOLD)



def duplicat():
     wb=load_workbook('save.xlsx')
     ws = wb.active
     ws.title='save'
     for i in range(1,ws.max_row + 1):
          alfa=ws.cell(row=i,column=2).value
          if str(SessionID.get()) == str(alfa):
               messagebox.showerror('DUPLICAT: error','this voice is DONE')
          else:
               voic=Label(root,text="this is ok :)))",font=10000)
               voic.pack(pady=0,padx=0,side=RIGHT,)
               voic.place(x=1150,y=15)
     wb.save(filename ="save.xlsx")
     wb.close
     #if str(SessionID.get()) not in voices:
          #voic=Label(root,text="this is ok :)))",font=10000)
          #voic.pack(pady=0,padx=0,side=RIGHT,)
          #voices.append(str(SessionID.get()))

     #else:
          #messagebox.showerror('DUPLICAT: error','this voice is DONE')
user=[]
def login():
     wb=load_workbook('name.xlsx')
     ws = wb.active
     ws.title='name'
     for i in range(1,ws.max_row + 1):
          alfa=ws.cell(row=i,column=5).value
          if str(username.get()) == str(alfa):
               beta=ws.cell(row=i,column=6).value
               if str(password.get())==str(beta):
                    log=Label(root,text="ok")
                    log.pack()
                    log.place(x=430,y=20)
                    use=(ws.cell(row=i,column=4).value)
                    us=Label(root,text=use)
                    us.pack(pady=1,padx=2)
                    us.place(x=240,y=45)
                    user.append(use)
               else:
                    messagebox.showerror('user and pass is incorrect !!!')
     wb.save(filename ="name.xlsx")
     wb.close
     
username=Label(root,text="username*")
username.pack()
username.place(x=180,y=5)
username=Entry(root)
username.pack()
username.place(x=240,y=5)

password=Label(root,text="password*")
password.pack()
password.place(x=180,y=30)
password=Entry(root,show='*')
password.pack()
password.place(x=240,y=30)

login=Button(root,text='login'
               ,command=login
               ,width=5
               ,height=2)
login.pack()
login.place(x=380,y=10)







SessionID=Label(root,text="Session ID",font=fonti)
SessionID.pack(ipadx=50,ipady=4)
SessionID.place(x=570,y=15)
SessionID=Entry(root,bd=8,font=fonti)
SessionID.pack(padx=5,pady=5)
SessionID.place(x=670,y=10)
cheked_dup=Button(root,text='cheked duplicat'
                  ,command=duplicat,width=17
                  ,height=2,font=fonti)
cheked_dup.pack()
cheked_dup.place(x=910,y=5)

Extension=Label(root,text="Extension",font=fonti)
Extension.pack(pady=1,padx=2)
Extension.place(x=210,y=100)
Extension=Entry(root,bd=5,font=fonti)
Extension.pack(pady=1,padx=2)
Extension.place(x=300,y=100)
L=Label(root,text="====>",font=fonti)
L.pack(pady=1,padx=2)
L.place(x=500,y=100)

namee=[]
def show():
     wb=load_workbook('name.xlsx')
     ws = wb.active
     ws.title='name'
     for i in range(1,ws.max_row + 1):
          alfa=ws.cell(row=i,column=1).value
          if str(Extension.get()) == str(alfa):
               ext=(ws.cell(row=i,column=2).value)
               namel=Label(root,text=ext,font=fonti)
               namel.pack(pady=1,padx=2)
               namel.place(x=670,y=100)
               namee.append(ext)
     wb.save(filename ="name.xlsx")
     wb.close
show_button=Button(root,text='show',command=show)
show_button.pack()
show_button.place(x=550,y=100)



name=Label(root,text="name",font=fonti)
name.pack(pady=1,padx=2)
name.place(x=600,y=100)
#name=Entry(root,bd=5,font=fonti)
#name.pack(pady=1,padx=2)
#name.place(x=670,y=100)

CalledNumber=Label(root,text="Called Number",font=fonti)
CalledNumber.pack(pady=1,padx=2)
CalledNumber.place(x=180,y=180)
CalledNumber=Entry(root,bd=5,font=fonti)
CalledNumber.pack(pady=1,padx=2)
CalledNumber.place(x=300,y=180)

calen = DateEntry(root,width=30,bg="darkblue",fg="white",year=2020)
calen.grid()
calen.place(x=300,y=143)

Date=Label(root,text="Date",font=fonti)
Date.pack(pady=1,padx=2)
Date.place(x=250,y=143)


dutration=Label(root,text="dutration",font=fonti)
dutration.pack(pady=1,padx=2)
dutration.place(x=570,y=143)
dutration=Entry(root,bd=5,font=fonti)
dutration.pack(pady=1,padx=2)
dutration.place(x=670,y=143)



karshenas=Label(root,text="توضیحات کارشناس",font=fonti)
karshenas.pack(pady=1,padx=2)
karshenas.place(x=890,y=100)
karshenas=Text(root,bd=5,font=fonti)
karshenas.pack(pady=5,padx=4)
karshenas.place(x=1000,y=100,width=400,height=240)


problems=[
'پیگیری جمع آوری',
'پیگیری برقراری',
'پیگیری تغییر پورت',
'پیگیری تغییر شماره',
'تغییر شماره',
'پیگیری رفع مشکل',
'استعلام خط/ ثبت نام',
'محموله ارسالی',
'نصب حضوری',
'دریافت نام کاربری و کلمه عبور',
'تغییر پهنای باند',
'درخواست جمع‌آوری',
'پیگیری عودت',
'اتفاقات شبکه',
' شکایت / درخواست SLA',
'باز نشدن پنل',
'درخواست اشکالات نرم افزاری',
'چراغ لینک خاموش',
'قطعی و وصلی',
'کندی سرعت',
' باز نشدن همه page  ها',
'نداشتن کانکشن',
'Ping time بالا',
'نویز/ چک کردن داخلی',
'تنظیمات Wirless',
'باز نشدن page خاص',
'عدم اتصال وایرلس',
'Error 651& 678',
'راهنمایی برای فعالسازی',
'Error 691',
'بازی آنلاین',
'کم شدن حجم بیشتر از میزان مصرف شده',
'کانفیگ مودم',
'تست از سرخط',
'مشکل در پرداخت',
'عدم رعایت اتصال همزمان',
'عدم تطابق حوزه اتصال',
'تنظیمات dvr',
'سرویس vpn، وایرلس و uso',
'کندی شبکه های اجتماعی',
'راهنمایی جهت نصب کردن نرم افزار',
' تنظیم/ مشکلات IP استاتیک  ',
'مشکلات مودم',
'عودت وجه',
'دریافت فاکتور',
'تماس با مشترک',
' نرم افزار My Hiweb ',
'پنل (خرید/ فعالسازی)',
'پنل (آموزش)',
'پنل (ثبت اطلاعات)',
'پنل (اطلاعات در مورد سرویس)',
'اطلاعات در مورد سایت',
'تغییر شماره- فنی',
'کسب اطلاعات درباره وضعیت خط',
'نامشخص-قطع تماس',

]

problemskarshenas=[
'تلفظ اشتباه لغات انگلیسی',
'تلفظ اشتباه لغات انگلیسی >> توضیحات اضافه اشتباه',
'لحن نامناسب در مکالمه',
'عدم معرفی کارشناس',
'عدم توجه به حرف و مشکل مشترک / پریدن وسط حرف مشترک',
'جملات تابو',
'خمیازه کشیدن، خندیدن، Mute نکردن و صحبت با شخص دیگر',
'عدم توانایی در توجیه و قانع کردن مشترک',
'استعلام خط/ ثبت نام',
'عدم آشنایی با تنظیمات مودم/ سیستم یا نرم‌افزار خاص',
'تنظیمات IP/ DVR',
'توضیح شفاهی (نویز، کانفیگ و ...)',
'دادن اطلاعات اشتباه به مشترک در اثر بی توجهی به crm یا مشکل مشترک',
'اطلاعات در مورد سایت',
'اطلاعات در مورد سایت >> توضیحات اضافه اشتباه,',
'اطلاعات در مورد سایت >> توضیحات اضافه - ناقص',
'عدم راهنمایی براساس قوانین مرکز تماس',
'مشکلات پرداخت',
'کامل چک نکردن بخشی از فلوچارت فنی',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت رانژه',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت dormant در انتظار فعالسازی',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت dormant فعال',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت up فعالسازی',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت قطعی پی در پی',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت کندی سرعت',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت نداشتن کانکشن',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت no page',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت نویز',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت ping time بالا/ packet loss',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت باز نشدن سایت خاص',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت erorr 691',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت erorr 651',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت بررسی ارتباط wireless',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت قطعی wireless',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت باز نشدن صفحه مودم',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت ping مودم',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت مغایرت در حجم مصرفی',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت تحویل به مشترک',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت 4G LTE',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت تغییر پهنای باند',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت دورمنت شدن حین تماس',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت تغییر شماره -فنی',
'کامل چک نکردن بخشی از فلوچارت فنی >> فلوچارت مشکل درگاه بانک',
'فلوچارت رانژه',
'فلوچارت dormant در انتظار فعالسازی',
'فلوچارت dormant فعال',
'فلوچارت up فعالسازی',
'فلوچارت قطعی پی در پی',
'فلوچارت کندی سرعت',
'فلوچارت نداشتن کانکشن',
'فلوچارت no page',
'فلوچارت نویز',
'فلوچارت ping time بالا/ packet loss',
'فلوچارت باز نشدن سایت خاص',
'فلوچارت erorr 691',
'فلوچارت erorr 651',
'فلوچارت بررسی ارتباط wireless',
'فلوچارت قطعی wireless',
'فلوچارت باز نشدن صفحه مودم',
'فلوچارت ping مودم',
'فلوچارت مغایرت در حجم مصرفی',
'فلوچارت تحویل به مشترک',
'فلوچارت 4G LTE',
'فلوچارت نصب حضوری',
'فلوچارت تغییر شماره',
'فلوچارت جمع‌آوری- عدم تلاش',
'فلوچارت عودت وجه',
'فلوچارت فن‌آوا',
'راهنمای اشتباه در مورد sal/شکایت',
'محموله ارسالی,',
'ثبت پیگیری ناقص',
'ثبت پیگیری اشتباه',
'عدم ثبت پیگیری/ درخواست',
'بستن درخواست',
'عدم بستن درخواست',
'متن آماده نامناسب',
'عدم تماس با مشترک',
'شکستن تماس',
'توضیحات اضافه بی مورد به صورت داوطلبانه',
'اتفاقات شبکه',
'اطلاع رسانی ناقص',
'عامیانه صحبت کردن',
'QC-فروش',
'بی احترامی به مشترک',
'راهنمایی ناقص',
'پنل (خرید/ فعالسازی)',
'پنل (خرید/ فعالسازی) >> توضیحات اضافه اشتباه',
'پنل (خرید/ فعالسازی) >> توضیحات اضافه - ناقص',
'پنل (آموزش)',
'پنل (آموزش) >> توضیحات اضافه اشتباه,',
'پنل (آموزش) >> توضیحات اضافه - ناقص',
'پنل (ثبت اطلاعات)',
',پنل (ثبت اطلاعات) >> توضیحات اضافه اشتباه',
'پنل (ثبت اطلاعات) >> توضیحات اضافه - ناقص',
'پنل (اطلاعات در مورد سرویس)',
'پنل (اطلاعات در مورد سرویس) >> توضیحات اضافه اشتباه',
'پنل (اطلاعات در مورد سرویس) >> توضیحات اضافه - ناقص',
'کامل چک نکردن بخشی از فلوچارت قوانین',
'کامل چک نکردن بخشی از فلوچارت قوانین >> فلوچارت نصب حضوری',
'کامل چک نکردن بخشی از فلوچارت قوانین >> فلوچارت تغییر شماره',
'کامل چک نکردن بخشی از فلوچارت قوانین >> فلوچارت ارسال فاکتور',
'کامل چک نکردن بخشی از فلوچارت قوانین >> فلوچارت جمع‌آوری- عدم تلاش',
'کامل چک نکردن بخشی از فلوچارت قوانین >> فلوچارت عودت وجه',
'کامل چک نکردن بخشی از فلوچارت قوانین >> فلوچارت فن‌آوا',
'کامل چک نکردن بخشی از فلوچارت قوانین >> فلوچارت آپلود مدارک',
'کامل چک نکردن بخشی از فلوچارت قوانین >> فلوچارت ثبت فیش',
'کامل چک نکردن بخشی از فلوچارت قوانین >> فلوچارت اشکالات نرم افزاری',
'کامل چک نکردن بخشی از فلوچارت قوانین >> فلوچارت ابطال سرویس',
'فلوچارت ارسال فاکتور',
'آموزش',
'مغایرت نظرسنجی',
'مغایرت نظرسنجی >> مغایرت نظرسنجی-لایه2',
'مغایرت نظرسنجی >> ثبت ناقص نظرسنجی-لایه2',
'مغایرت نظرسنجی >> مغایرت نظرسنجی-لایه1',
'فلوچارت تغییر پهنای باند',
'فلوچارت دورمنت شدن حین تماس',
'فلوچارت تغییر شماره -فنی',
'تماس مجدد',
'فلوچارت مشکل درگاه بانک',
'فلوچارت آپلود مدارک',
'فلوچارت ثبت فیش',
'فلوچارت ابطال سرویس',
'فلوچارت اشکالات نرم افزاری',
'قطع تماس حین مکالمه',
'عدم تلاش جهت حل مشکل',

]

grad=[
     '1','1.5',
     '2','2.5',
     '3','3.5',
     '4'
     ]

estandard=[
'عامیانه و غیر رسمی صحبت کردن',
'عدم معرفی کارشناس',
'بی حوصله صحبت کردن',
'عدم توانایی در توجیه و قانع کردن مشترک',
'عدم توجه به حرف و مشکل مشترک',
'کارشناس بگوید همه مشترکین این مشکل را دارند یا اختلال، سراسری است',
'در مواقع قطعی کارشناس اعلام می کند "ما خودمون هم نمیدونستیم که قطع شدیم از مشترک ها متوجه شدیم"',
'همکارانی که دسترسی دارند الان در شرکت حضور ندارند',
'استفاده از کلمات یا جملات تابو',
'مشکل شما را باید از همکاران بپرسم.',
'جملاتی که باعث میشود مشترک حس کند همکاران در حال مچ گیری یکدیگر هستند',
'به مشترک اعلام شود ساعت نهارم است',
'واضح نبودن صدا',
'خمیازه کشیدن، خندیدن، Mute نکردن و صحبت با شخص دیگر',
'مواردی که کارشناس با شک و تردید جواب می دهد که ناشی از اطلاعات کم کارشناس است',
'راهنمایی براساس فلوچارت نباشد و مشکل مشترک حل نشود',
'راهنمایی براساس فلوچارت نباشد و مشکل مشترک در زمان نامناسب حل شود',
'عدم اطمینان از حل شدن مشکل مشترک و قطع کردن تلفن',
'توضیحات شفاهی',
'دادن اطلاعات اشتباه به مشترک در اثر بی توجهی به CRM یا مشکل مشترک',
'عدم راهنمایی براساس قوانین مرکز تماس',
'راهنمایی اشتباه و براساس آن ثبت پیگیری اشتباه که باعث گمراه شدن کارشناس بعدی شود',
'عدم تکمیل یا تغییر اطلاعات',
'عدم تلاش جهت نگهداشت مشترکی که قصد جمع آوری دارد',
'توضیحات اضافه بی مورد به صورت داوطلبانه',
'نیاز به باز کردن درخواست نیست و درخواست بی مورد اضافه شده و بسته نشده',
'در صورت حل نشدن یا عدم اطمینان از حل شدن \nمشکل مشترک باید درخواست باز بماند ولی کارشناس درخواست را می بندد',
'مشترک درخواست باز دارد ولی پیگیری ثبت نمی کنند و عدم توجه به CRM',
'کارشناس باید پیگیری و یا درخواست ثبت کند ولی ثبت نمی کند',
'عدم استفاده از متن آماده شروع و پایان مناسب در درخواستها',
'در صورتی که شیفت کاری به اتمام رسیده است و هنوز مکالمه با مشترک تمام نشده به هیچ عنوان نباید مشترک را به کارشناس دیگری ارجاع داد',
'درصورتیکه قرار است برای مشترک Selt گرفته شود و یا موردی پیگیری شود، به مشترک اعلان شود با او تماس گرفته می‌شود.',
'چک نکردن مشخصات مشترک (اسم یا نام کاربری یا شماره تلفن یا ...) با وی در ابتدای مکالمه که منجر به بروز مشکل می شود',
'مواردی که کارشناس به دلیل عصبانیت و یا هر مساله دیگری تلفن را قطع می کند',
'بی دلیل تماس قطع و یا hold شود',
'تلفظ اشتباه لغات انگلیسی',
'عدم برگرداندن پروفایل به حالت مناسب',
'هر گونه بازخواست، تمسخر، تحقیر، بی احترامی و یا توهین به مشترک',
'تبلیغ مثبت برای شرکت های همکار و ذکر مزایای آنها در مقایسه با شرکت های وب',
'راهنمایی براساس فلوچارت باشد اما ناقص انجام شود و مشکل حل نشود',
'توضیحات پنل کاربری داده شود ولی در مورد سایر موارد(اپ اندروید، ربات تلگرام و لینک آموزشی) به مشترک اطلاع داده نشود',
'مشکل مشترک حل شد اما توضیحات پنل و اطلاع رسانی سایر موارد (اپ اندروید، ربات تلگرام و لینک آموزشی) انجام نشد',
'داکیومنت های لازم (با اینکه که قبلا برای مشترک ارسال نشده بود) برای مشترک ارسال نشد',
'عدم اطلاع رسانی در مورد ارسال داکیومنت',
'در صورت حل شدن مشکل کارشناس باید پیگیری ثبت کند و درخواست را ببندد (پیگیری ثبت نشده و درخواست بسته نشده است)',
'در صورت حل شدن مشکل کارشناس باید پیگیری ثبت کند و درخواست را ببندد (پیگیری ثبت شده ولی درخواست بسته نشده است)',
'ثبت اشتباه اطلاعات مشترک',
'مواردی که توسط مشترک قابل انجام است ولی به دلیل عدم توانایی درخواست دارد توسط کارشناس انجام شود اما کارشناس اقدام لازم را انجام نمیدهد',
'در زمان قطعی یا اختلال کارشناس براساس پرامپت های IVR (موجود در فلوچارت) راهنمایی نکند',
'کارشناس در تماس اول توضیحات اشتباه یا ناقص به مشترک ارائه کند ولی مجدد با مشترک تماس بگیرد و توضیحات را تصحیح کند',
'در طی مراحل فلوچارت کارشناس تنظیمات مودم را تغییر میدهد ولی در انتهای مکالمه تنظیمات را به حالت قبل برنمیگرداند ویا با رضایت مشترک تحویل نمیدهد',
'در صورتی که مشترک درخواست باز دارد و برای مشکل دیگری تماس گرفته که نیاز به ثبت پیگیری یا درخواست دارد و اعلام میکند مشکل اولیه حل شده، کارشناس درخواست را نبندد و درخواست جدید با متن مناسب ایجاد نکند',
'کلیه مواردی که کارشناس باید برای پیگیری به ارشد یا بخش دیگری ارجاع دهد',
'راهنمایی براساس فلوچارت باشد اما ناقص انجام شود و مشکل حل شود',
'مشترک اصرار دارد کارشناس منتظر بماند و تلفن را قطع نکند ولی کارشناس تماس را قطع کند',
'توضیحات پنل کاربری داده نشود و سایر موارد اطلاع رسانی نشود',
'توضیحات پنل کاربری داده نشود و سایر موارد اطلاع رسانی شود',
'زمانی که مشترک به مودم و سیستم دسترسی ندارد توضیحات شفاهی داده نشود و سایر موارد اطلاع رسانی نشود',
'زمانی که مشترک به مودم و سیستم دسترسی ندارد توضیحات شفاهی داده نشود و سایر موارد اطلاع رسانی شود',
'در صورتی که در مکالمه در مورد خرید صحبت میشود، کارشناس در مورد آشنایی به نحوه خرید از مشترک سوال نپرسد',
"به دلیل راهنمایی اشتباه مشکل حل نمی شود",
'"به دلیل راهنمایی ناقص مشکل حل نمی شود"',
'مکالمه قطع شود و کارشناس با مشترک تماس نگیرد یا درخواست و پیگیری ثبت نکند',
'پیگیری مشکل مشترک یا ثبت پیگیری/ درخواست با تاخیر بیش از 30 دقیقه',
'پیگیری یا درخواست، ناقص یا بخشی از آن اشتباه ثبت می شود',
'در مراحلی از فلوچارت که مشترک قادر به چک کردن موارد نیست و کارشناس براساس مسیر (بله یا خیر) مشخص شده ادامه را چک میکند، کارشناس انتهای مکالمه مشترک را جهت بررسی مورد چک نشده راهنمایی نکند و مورد چک نش ...',
'اگر ایمیل / شماره موبایل مشترک ثبت نیست و تماس از خطوط ثبت شده نیست کارشناس برای ثبت ایمیل / شماره موبایل از طریق پنل کاربری راهنمایی نکند',
'اگر ایمیل / شماره موبایل مشترک ثبت نیست و تماس از خط ADSL یا موبایل است کارشناس ایمیل / شماره موبایل مشترک را دریافت و ثبت نکند',
'کارشناس تمام تلاش خود را جهت حل مشکل مشترک انجام نداده است',

     ]

nazarsanji=['0','1']




clicked1=StringVar(root)
clicked1.set('مشکل 1 کاربر')
problem1=ttk.Combobox(root,textvariable=clicked1,state="readonly",values=problems,width=25)
problem1.pack()
problem1.config(font=fonti)
problem1.place(x=280,y=220)

clicked12=StringVar()
clicked12.set('مشکل 2 کاربر')
problem12=ttk.Combobox(root,textvariable=clicked12,state="readonly",values=problems,width=25)
problem12.pack()
problem12.config(font=fonti)
problem12.place(x=280,y=260)


clicked13=StringVar()
clicked13.set('مشکل 3 کاربر')
problem13=ttk.Combobox(root,textvariable=clicked13,state="readonly",values=problems,width=25)
problem13.pack()
problem13.config(font=fonti)
problem13.place(x=280,y=300)


clicked14=StringVar()
clicked14.set('مشکل 4 کاربر')
problem14=ttk.Combobox(root,textvariable=clicked14,state="readonly",values=problems,width=25)
problem14.pack()
problem14.config(font=fonti)
problem14.place(x=280,y=340)

clicked2=StringVar()
clicked2.set('مشکل 1 کارشناس')
problemskarshenas1=ttk.Combobox(root,textvariable=clicked2,state="readonly",values=problemskarshenas,
                                width=48)
problemskarshenas1.pack()
problemskarshenas1.config(font=fonti)
problemskarshenas1.place(x=570,y=220)

clicked22=StringVar()
clicked22.set('مشکل 2 کارشناس')
problemskarshenas12=ttk.Combobox(root,textvariable=clicked22,state="readonly",values=problemskarshenas,
                                 width=48)
problemskarshenas12.pack()
problemskarshenas12.config(font=fonti)
problemskarshenas12.place(x=570,y=260)

clicked23=StringVar()
clicked23.set('مشکل 3 کارشناس')
problemskarshenas13=ttk.Combobox(root,textvariable=clicked23,state="readonly",values=problemskarshenas,
                                 width=48)
problemskarshenas13.pack()
problemskarshenas13.config(font=fonti)
problemskarshenas13.place(x=570,y=300)

clicked24=StringVar()
clicked24.set('مشکل 4 کارشناس')
problemskarshenas14=ttk.Combobox(root,textvariable=clicked24,state="readonly",values=problemskarshenas,
                                 width=48)
problemskarshenas14.winfo_reqwidth()
problemskarshenas14.pack()
problemskarshenas14.config(font=fonti)
problemskarshenas14.place(x=570,y=340)

clicked3=StringVar()
clicked3.set('استاندارد')
estandard1=ttk.Combobox(root,textvariable=clicked3,state="readonly",
                        values=estandard,width=100)
estandard1.winfo_width()
estandard1.pack(ipadx=100, ipady=100)
estandard1.config(font=fonti)
estandard1.place(x=280,y=360)

clicked4=StringVar()
clicked4.set('لحن')
grad1=ttk.Combobox(root,textvariable=clicked4,state="readonly",values=grad)
grad1.winfo_reqwidth()
grad1.pack()
grad1.config(font=fonti)
grad1.place(x=1100,y=400)

clicked42=StringVar()
clicked42.set('راهنمايي مناسب جهت حل مشکل')
grad12=ttk.Combobox(root,textvariable=clicked42,state="readonly",values=grad)
grad12.pack()
grad12.config(font=fonti)
grad12.place(x=1100,y=440)

clicked43=StringVar()
clicked43.set('پيگيري')
grad13=ttk.Combobox(root,textvariable=clicked43,state="readonly",values=grad)
grad13.pack()
grad13.config(font=fonti)
grad13.place(x=1100,y=480)


  
clicked5=StringVar()
clicked5.set('مغايرت  نظر سنجي لايه1')
nazarsanji1=ttk.Combobox(root,textvariable=clicked5,state="readonly",values=nazarsanji)
nazarsanji1.pack()
nazarsanji1.config(font=fonti)
nazarsanji1.place(x=670,y=400)

clicked52=StringVar()
clicked52.set('مغايرت نظرسنجي لايه2')
nazarsanji12=ttk.Combobox(root,textvariable=clicked52,state="readonly",values=nazarsanji)
nazarsanji12.pack()
nazarsanji12.config(font=fonti)
nazarsanji12.place(x=670,y=440)

clicked53=StringVar()
clicked53.set('ثبت ناقص نظرسنجي')
nazarsanji13=ttk.Combobox(root,textvariable=clicked53,state="readonly",values=nazarsanji)
nazarsanji13.pack()
nazarsanji13.config(font=fonti)
nazarsanji13.place(x=670,y=480)



book=load_workbook("save.xlsx")
savee=book.active
savee.title='save'
cal=savee['A1']
if cal.value is None:
     x=1
     savee.cell(row=1,column=1).value=int(x)

y =savee['A1']
y=y.internal_value
book.save(filename ="save.xlsx")
book.close




def donn():
     global x
     global y
     book=load_workbook("likeofexcel1.xlsx")
     sheett=book.active
     sheett.title='QC'
     #if x==1:

     a=['session','Extension','name','CalledNumber','Date','dutration',
          "توضیحات کارشناس",'مشکل 1 کاربر','مشکل 1 کاربر',
          'مشکل 2 کاربر',
          'مشکل 3 کاربر',
          'مشکل 4 کاربر',
          'مشکل 1 کارشناس',
          'مشکل 2 کارشناس',
          'مشکل 3 کارشناس',
          'مشکل 4 کارشناس',
          'استاندارد',
          'لحن',
          'راهنمايي مناسب جهت حل مشکل',
          'پيگيري',
          'مغايرت  نظر سنجي لايه1',
          'مغايرت نظرسنجي لايه2',
          'ثبت ناقص نظرسنجي','ارزياب']
     for j in range(1,(len(a))+1):
          sheett.cell(row=1, column=j).value =a[j-1]
               
     #elif x>0:
     b=[SessionID.get(),Extension.get(),namee[0],CalledNumber.get(),calen.get(),
        dutration.get(),karshenas.get("1.0",END),clicked1.get(),clicked12.get(),clicked13.get(),
        clicked14.get(),clicked2.get(),clicked22.get(),clicked23.get(),clicked24.get(),
        clicked3.get(),clicked4.get(),clicked42.get(),clicked43.get(),clicked5.get(),
        clicked52.get(),clicked53.get(),user[0]]
     for j in range(1,(len(b))+1):
          sheett.cell(row=y+1, column=j).value =b[j-1]
               
     book.save(filename ="likeofexcel1.xlsx")
     book.close
     y+=1
     book=load_workbook("save.xlsx")
     savee=book.active
     savee.title='save'
     cal=savee['A1']
     savee.cell(row=1,column=1).value=int(y)
     book.save(filename ="save.xlsx")
     book.close

     
     wb=load_workbook('save.xlsx')
     ws = wb.active
     ws.title='save'
     cal=ws.cell(row=y+1,column=2)
     cal.value=SessionID.get()
     wb.save(filename ="save.xlsx")
     wb.close
     SessionID.delete(first=0,last=100)
     karshenas.delete(1.0,END)
     Extension.delete(first=0,last=100)
     CalledNumber.delete(first=0,last=100)
     dutration.delete(first=0,last=100)
     na=Label(root,text='                               ',font=fonti)
     na.pack(pady=10,padx=6)
     na.place(x=1150,y=15)
     namel=Label(root,text='                               ',font=fonti)
     namel.pack(pady=10,padx=6)
     namel.place(x=670,y=100)
     clicked1.set('مشکل 1 کاربر')
     clicked12.set('مشکل 2 کاربر')
     clicked13.set('مشکل 3 کاربر')
     clicked14.set('مشکل 4 کاربر')
     clicked2.set('مشکل 1 کارشناس')
     clicked22.set('مشکل 2 کارشناس')
     clicked23.set('مشکل 3 کارشناس')
     clicked24.set('مشکل 4 کارشناس')
     clicked3.set('استاندارد')
     clicked4.set('لحن')
     clicked42.set('راهنمايي مناسب جهت حل مشکل')
     clicked43.set('پيگيري')
     clicked5.set('مغايرت  نظر سنجي لايه1')
     clicked52.set('مغايرت نظرسنجي لايه2')
     clicked53.set('ثبت ناقص نظرسنجي')


done_button=Button(root,text='done',command=donn)
done_button.pack()
done_button.place(x=910,y=600)

def donothing():
     filewin = Toplevel(root)
     button = Button(filewin, text="Do nothing button")
     button.pack()
     menubar = Menu(root)
     filemenu = Menu(menubar, tearoff=0)#thisis menubar
     filemenu.add_command(label="New", command=donothing)
     filemenu.add_command(label="Open", command=donothing)
     filemenu.add_command(label="Save", command=donothing)
     filemenu.add_command(label="Save as...", command=donothing)
     filemenu.add_command(label="Close", command=donothing)
     filemenu.add_separator()#this is line
     filemenu.add_command(label="Exit", command=root.quit)
     menubar.add_cascade(label="File", menu=filemenu)
     editmenu = Menu(menubar, tearoff=0)
     editmenu.add_command(label="Undo", command=donothing)
     editmenu.add_separator()
     editmenu.add_command(label="Cut", command=donothing)
     editmenu.add_command(label="Copy", command=donothing)
     editmenu.add_command(label="Paste", command=donothing)
     editmenu.add_command(label="Delete", command=donothing)
     editmenu.add_command(label="Select All", command=donothing)
     menubar.add_cascade(label="Edit", menu=editmenu)
     helpmenu = Menu(menubar, tearoff=0)
     helpmenu.add_command(label="Help Index", command=donothing)
     helpmenu.add_command(label="About...", command=donothing)
     menubar.add_cascade(label="Help", menu=helpmenu)
     root.config(menu=menubar)
     
root.mainloop()
'''
calen=Calendar(root,selectmode='day',year=2020,month=5,day=22)
calen.pack()
calen.place(x=220,y=400)
def grab_date():
     datelabel.config(text=''+calen.get_date())


datelabel=Label(root,text='',font=fonti)
datelabel.pack(pady=1,padx=2)
datelabel.place(x=300,y=143)
print(datelabel)

datebutton=Button(root,text="Get date",command=grab_date)
datebutton.pack()
'''
#Date=Entry(root,bd=5,font=fonti)
#Date.pack(pady=1,padx=2)
#Date.place(x=300,y=143)

#nameqc=Label(root,text="ارزياب",font=fonti)
#nameqc.pack(pady=1,padx=2)
#nameqc.place(x=570,y=600)
#nameqc=Entry(root,bd=5)
#nameqc.pack(pady=1,padx=2)
#nameqc.place(x=670,y=600)

'''
     global x
     print(10)
     book=Workbook()
     sheett=book.active
     sheett.title='QC'
     rowss = (
          ('SessionID',
          'Extension',
          'name'),
         (SessionID.get(), Extension.get(), name.get()),
          )
     
          book.save("hiwebQC.xlsx")

     else:
          print(55)
          book=Workbook()
          sheett=book.active
          sheett.title='QC'
          b=x+2
          sheett.insert_rows(b)
          sheett.iter_rows(min_row=b, max_row=None, min_col=None, max_col=None, values_only=False)
          rowss=(SessionID.get(), Extension.get(), name.get())
          val=[SessionID.get(), Extension.get(), name.get()]

          for val in val:
               sheett.cell(row=b, column=column).value = val
          sheett.append(rowss)

          book.save("hiwebQC.xlsx")
def donn():
     rr=open('hiweb2.csv','r+')
     ff=['SessionID',
          'Extension',
          'name',
          'CalledNumber',
          'Date']
     writer=csv.DictWriter(rr,fieldnames=ff)
     writer.writeheader()
     reader=csv.reader(rr)
     for row in range(100):
          #if row[0] in (None, ""):

              writer.writerow({'SessionID':[SessionID.get()],
                              'Extension':[Extension.get()],
                              'name':[name.get()],
                              'CalledNumber':[CalledNumber.get()],
                              'Date':[Date.get()]})
i=1

def donn():
     global i  
     result=pd.DataFrame({
          'SessionID':[SessionID.get()],
          'Extension':[Extension.get()],
          'name':[name.get()],
          'CalledNumber':[CalledNumber.get()],
          'Date':[Date.get()],
          'dutration':[dutration.get()],
          'karshenas':[karshenas.get()],
          'problem1':[clicked1.get()],
          'problem2':[clicked12.get()],
          'problem3':[clicked13.get()],
          'problem4':[clicked14.get()],
          'problemskarshenas1':[clicked2.get()],
          'problemskarshenas12':[clicked22.get()],
          'problemskarshenas13':[clicked23.get()],
          'problemskarshenas14':[clicked24.get()],
          'estandard':[clicked3.get()],
          'grad1':[clicked4.get()],
          'grad2':[clicked42.get()],
          'grad3':[clicked43.get()],
          'nazarsanji1':[clicked5.get()],
          'nazarsanji12':[clicked52.get()],
          'nazarsanji13':[clicked53.get()],
     })
     xlwriter=pd.ExcelWriter('hiwebQC.xlsx')
     result.to_excel(xlwriter,sheet_name='QC',index=True,)
     result=result.iloc[0:i,:]
     result =result.append(result, ignore_index=True,verify_integrity=True)

     xlwriter.close()
     i+=1
'''     

